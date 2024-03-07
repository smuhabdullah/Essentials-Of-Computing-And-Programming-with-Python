from tkinter import *
from tkinter import ttk

import customtkinter
from CTkListbox import *
from CTkMessagebox import CTkMessagebox
from customtkinter import *
from PIL import Image
import tkinter 
import openpyxl
import datetime


# I will give a path where my product file is
path = "data/products.xlsx"
#now I will open the file with openpyxl module
workbook = openpyxl.load_workbook(path)
#now I will select the sheet I want to use
sheet = workbook.active 




#images
img = customtkinter.CTkImage(Image.open('images/logo1.png'),size=(170,170))





# Set the appearance of the custom Tkinter to dark
customtkinter.set_appearance_mode("dark")
customtkinter.set_default_color_theme("dark-blue")  
# Themes: "blue" (standard), "green", "dark-blue"

# now i will Create the main window (root)
root = customtkinter.CTk()




# now I will use .geometry to set a size of the window and
#I will give the width and height variable as arguments

root.geometry("1500x900")

#root.state("iconic")


#now I will set the title of the window

root.title('POS System')


# Add an icon to the window (uncomment and provide the correct path to your icon file)
#root.iconbitmap("images/logo1.ico")



#creating the first function

#here I am defined the function of add to cart

def add_to_cart():

 #first we will we will take product and quantity from the entry boxes with .get
  product = enter_product.get().strip()
  # where I will use try and except method to produce an error if you user enters a string
  
  quantity = int (enter_quantity.get())
  
  #now I will loop through the rows of the sheet and add the data to the dictionary

  for row in sheet.iter_rows():


    #now I will get the product name price and quantity from the excel file

    product_name =row[0].value
    product_price =row[1].value
    product_stock =row[2].value


    #here i will use if conditíons to check the product

    if product == product_name:

      # now will check if product stock is available

      if product_stock >= quantity:

        #using another if so that it only works when stock is available
        #now i will total the price of the product_stock

        total_price = product_price * quantity

        #now I will format the product detail as a string
        product_details=f"{product_name}            ×            {quantity}             ×          {product_price}               =                     {total_price}"
        #now I will add this formatted string into my cart list box
        cart_listbox.insert("end",product_details)
        #now I will update the stock in my excel file
        sheet.cell(row, 2,product_stock - quantity)
        #now I will save the changes in my excel file and break the loop
        workbook.save(path)

        cart_listbox.insert("end", product_details)
        #using .delete to delete text from entry boxes
        enter_product.delete(0, "end")
        enter_quantity.delete(0, "end")
        break
    #give same code by using now I will create an else statement that will show an error if product stock is not enough
    #using CTkMessagebox module i imported to show an error
    
    
      else:
        CTkMessagebox(title="Error", message="Product is out of stock", icon="warning")
      enter_product.delete(0, "end")
      enter_quantity.delete(0, "end")
      break
    
    elif product == "" and quantity == "":
      CTkMessagebox(title="Error", message="Please enter Product Name and password", icon="warning")


    elif product == "":
      CTkMessagebox(title="Error", message="Please enter Product Name", icon="warning")

    elif quantity == "":
      CTkMessagebox(title="Error", message="Please enter Quantity", icon="warning")
    
    
  else:
    #Error when Product is not found
    CTkMessagebox(title="Error", message="Product Not Found", icon="warning")
    enter_product.delete(0, "end")
    enter_quantity.delete(0, "end")

#Add to cart function done----------------------


#now i am writing exit_program func

#exit program function


def exit_program():
    # Show some retry/cancel warnings
  msg = CTkMessagebox(master=root,
    title="Are You Sure!",
  message="Do you want exit?",
  icon="warning",
  option_1="Yes",
  option_2= "No")
  yes_or_no = msg.get()
  if yes_or_no == "Yes":
      root.destroy()
  else:
      print("press yes to exit")


#exit_program function done 


#now calculate_total function

def calculate_total():
  #first make a total amount var and gave it 0 
  total_amount = 0 
  #now i will loop through the cart
  for item in cart_listbox.get(0,"end"):

    #where I will create a variable called item parts and I will use items dot split to convert product detail strain into different strings its like it will create an list with different strings

    product, price, quantity = item.split(" ")
        # Convert the price and quantity to float and int respectively
    price = float(price)
    quantity = int(quantity)
        # Multiply the price and quantity and add it to the total
    total_amount += price * quantity
    # Display the total in a label
    total_label.config(text=f"Total: {total}")


#calculate_total function done

#login function

def login():
  #first I will get the username and password from the entry boxes
  
  username = enter_username.get()
  password = enter_password.get()
  
  #now I will check if the username and password are correct
  if username == "admin" and password == "admin":
    #if correct creating a new top level window  in root window with argument root
    admin = customtkinter.CTkToplevel(root)
    
    admin.geometry("800x600")
    admin.maxsize(800,600)
    admin.title("Admin Page")
    #this WM transient command will put top level window in front of my main window
    #thank to stackoverflow i found this command
    admin.wm_transient(root)
  

#now I WILL configured the rows and columns in my admin window
    admin.columnconfigure(0,weight =2,uniform="a")
    admin.columnconfigure(1,weight =3,uniform="a")
    admin.rowconfigure(0,weight =1,uniform="a")
    admin.rowconfigure(1,weight =3,uniform="a")



  #first i will create the admin heading frame in admin window

    admin_heading_frame = customtkinter.CTkFrame(admin, width=100, height=100)
     #now i will place the heading frame with grid
    admin_heading_frame.grid(row=0, column=0, columnspan=3,padx=3, sticky=tkinter.EW)

     #now i will give rows and columns to  the admin heading frame

    admin_heading_frame.columnconfigure(1,weight =1,uniform="a")
    admin_heading_frame.rowconfigure(1,weight =1,uniform="a")

    admin_heading_Label= customtkinter.CTkLabel(admin_heading_frame ,text='HELLO ADMIN',
      font=('times new roman',45,'bold'),
      text_color='yellow',
      image=img, # i imported this img at the start of the code
      compound=LEFT)


    admin_heading_Label.grid(row=0, column=0, columnspan=2,padx=5, sticky=tkinter.NSEW)






    product_frame = customtkinter.CTkFrame(admin, width=200, height=100)

    # placing my frames in admin window with .grid

    product_frame.grid(row=1, column=0, columnspan=1,padx=5, sticky=tkinter.NSEW)
    
    product_frame.columnconfigure(0,weight =1,uniform="a")
    product_frame.rowconfigure(0,weight =1,uniform="a")
    product_frame.rowconfigure(1,weight =1,uniform="a")
    product_frame.rowconfigure(2,weight =1,uniform="a")
    product_frame.rowconfigure(3,weight =1,uniform="a")
    




#now I will create entry boxes and buttons
    #buttons
    


#entry boxes

    enter_product_name = customtkinter.CTkEntry(product_frame,
      placeholder_text= "Enter product name",
      width=500,
      height=100,
      font=("helvetica",24),
      text_color="#F5DD90",
      placeholder_text_color="yellow",
      corner_radius=200)


    enter_product_quantity = customtkinter.CTkEntry(product_frame,
      placeholder_text= "Enter product quantity",
      width=500,
      height=60,
      font=("helvetica",24),
      text_color="#F5DD90",
      placeholder_text_color="yellow",
      corner_radius=200)




    enter_product_price = customtkinter.CTkEntry(product_frame,
      placeholder_text= "Enter Product Price",
      width=500,
      height=60,
      font=("helvetica",24),
      text_color="#F5DD90",
      placeholder_text_color="yellow",
      corner_radius=200)

    
    save_product = customtkinter.CTkButton(product_frame, text ="Save Product",
      width =120,
      height =60,
      font=("helvetica",19),
      text_color="black",
      fg_color="yellow",
      hover_color="#c2b84e",
      corner_radius=200,
                                          )
    



    # placing the widgets in the product_add_frame
    enter_product_name.grid(row=0, column=0, padx=5, pady=5,)
    enter_product_price.grid(row=1, column=0, padx=5, pady=5,)
    enter_product_quantity.grid(row=2, column=0, padx=5, pady=5,)
    save_product.grid(row=3, column=0, padx=5, pady=5,)







    
    exel_file_frame = customtkinter.CTkFrame(admin, width=100, height=100)







    # this excel file frame will contain my excel file and show it in my admin window

    exel_file_frame.grid(row=1, column=1,padx=5, sticky=tkinter.NSEW)





  # setting up the function that will run when the save button is clicked

    # configuring excel file frame

    exel_file_frame.columnconfigure(0, weight=1)
    exel_file_frame.columnconfigure(1, weight=3)
    exel_file_frame.columnconfigure(2, weight=1)
    exel_file_frame.rowconfigure(0, weight=1)



    # so here I will use TTK frame from TTK module
    tree_frame =ttk.Frame(exel_file_frame, width=100, height=500)
    # will it a grid
    tree_frame.grid(row=0, column=0, columnspan=2, padx=5, pady=5, sticky=tkinter.NSEW)


    tree_frame.columnconfigure(0, weight=1)
    tree_frame.rowconfigure(0, weight=1)

    # now i will create a treeview in tree frame
    

    
  # here I will create a scroll bar to connect to our tree view to scroll
    
    tree_scroll = ttk.Scrollbar(tree_frame,)
    # here I will pack it  to the right side 
    tree_scroll.grid(row=0, column=1, sticky=tkinter.NSEW)

# now I will create a tuple named colm and i will put names of my columns in excel file
    colm = ("name","price","stock")

    
    #now I will create a treeview with the columns i created
    
    tree_view = ttk.Treeview(tree_frame,
      show="headings",
      yscrollcommand=tree_scroll.set,
      columns=colm, height=13 )

    
    # now I will configure the size of the columns of my tree view

    
    
    tree_view.column("name", width=50)
    tree_view.column("price", width=35)
    tree_view.column("stock", width=25)
    




    
    # now i will pack it

    tree_view.grid(row=0, column=0, padx=5, pady=5, sticky=tkinter.NSEW)

    tree_scroll.config(command=tree_view.yview)



    # now I will use openpyxl modular imported before

      # now I will make a function to load data
    def load_data():
    
  
    # now I will create a list with tuples of all the data

      list_values = list(sheet.values)
      print(list_values)
    #now I will loop through the list and add the data to the treeview
      for col_name in list_values[0]:
    # now I will specify the name of my heading in my tree view
        tree_view.heading(col_name, text=col_name)
        for value_tuple in list_values[1:]:
          tree_view.insert('', tkinter.END, values=value_tuple)


    # now I will loop through the list values and add the data to the treeview

    

    load_data()


      # let's add two nested functions

    def save_product():

      #first we will get the product name, quantity and price from the entry boxes
      product_name = enter_product_name.get()
      product_stock = enter_product_quantity.get()
      product_price = enter_product_price.get()

    # now I will load the excel file same as  I did before in load data function

    # right now I am inserting data into Excel sheet
      path = "data/products.xlsx"
      workbook = openpyxl.load_workbook(path)
      #now I will select the sheet I want to use
      sheet = workbook.active
  
      # I will give the row values
      row_values =["name", "price",'stock']

      # now I will append these ROws into sheets
  
      sheet.append(row_values)
      # now I will save the workbook
      workbook.save(path)
  
      # adding this data into tree view
  
      tree_view.insert("", tkinter.END, values=row_values)


      #now i will create save button





  elif username == "" and password == "":
      CTkMessagebox(title="Error", message="Please enter username and password", icon="warning")


  elif username == "":
    CTkMessagebox(title="Error", message="Please enter username", icon="warning")

  elif password == "":
    CTkMessagebox(title="Error", message="Please enter password", icon="warning")

  else:
    CTkMessagebox(title="Error", message="Invalid username or password", icon="warning")


# login function then now I will work on print bill function


def print_bill():
  
  # i will get the customer name from the customer_name_entry widget
  customers_name = customer_name.get()
   # i will get the customer number from the customer_number_entry widget
  customers_number = customer_number.get()
    # i will get the current date and time as a string
  date_time = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    # i will clear the bill_textbox widget
  bill_textbox.delete("1.0", "end")
    # now i will insert the customer name, number, and date and time into the bill_textbox widget
  bill_textbox.insert("end", f"Customer Name: {customer_name}\n")
  bill_textbox.insert("end", f"Customer Number: {customer_number}\n")
  bill_textbox.insert("end", f"Date and Time: {date_time}\n")
    # now i will insert a separator line into the bill_textbox widget
  bill_textbox.insert("end", "-" * 40 + "\n")
    # loop through the items in the cart listbox
  for item in cart_listbox.get(0, "end"):
        # now i will insert the item into the bill_textbox widget
    bill_textbox.insert("end", f"{item}\n")
    # now i will again insert a separator line into the bill_textbox widget
    bill_textbox.insert("end", "-" * 40 + "\n")
    # now i will get the total amount from the total_label widget
    total_amount = total_label.cget("text")
    # now i will insert the total amount into the bill_textbox widget
    bill_textbox.insert("end", f"{total_amount}\n")
    # now i will insert a thank you message into the bill_textbox widget
    bill_textbox.insert("end", "Thank you for shopping with us!\n")
    



def clear_cart():
  # clear the cart_listbox widget
    cart_listbox.delete(0, "end")
    # clear the total_label widget
    total_label.config(text="Total: 0")
  
    # show a success message
    CTkMessagebox(title="Success", message="Cart cleared successfully", icon="check")
    

  


#lets make heading menu i will include a logo ,name and login function
#first i will make a frame in which i will put widgets

heading_frame = customtkinter.CTkFrame(master=root, width=1100, height=500)
#now i will make a Username and Password entryboxes

enter_username= customtkinter.CTkEntry(heading_frame,
    placeholder_text= "Enter Username",
    width=460,
    height=60,
    font=("helvetica",24),
    text_color="#F5DD90",
    placeholder_text_color="yellow",
    corner_radius=200,)

enter_password = customtkinter.CTkEntry(heading_frame,
    placeholder_text= "Enter Password",
    width=460,
    height=60,
    font=("helvetica",24),
    text_color="#F5DD90",
    placeholder_text_color="yellow",
    corner_radius=200)

#i have entered the name of frame i created as the first argument  it is neccessary

#here i maked a login button ans at last set the command to login function i created earlier

login_button = customtkinter.CTkButton(heading_frame, text ="Login",
    width =120,
    height =60,
    font=("helvetica",19),
    text_color="black",
    fg_color="yellow",
    hover_color="#c2b84e",
    corner_radius=200,
    command=login)

#now i have created the main text label 

headingLabel= customtkinter.CTkLabel(heading_frame ,text='KAMAL PAINT PORTAL',
  font=('times new roman',50,'bold'),
  text_color='yellow',
  image=img, # i imported this img at the start of the code
  compound=LEFT)

  ##so there are two function that are used to place widgets .pack and .grid  so I will use . grid to place the widgets in frame and window and i will do it ate the end of the code





  #-----------------heading frame done----------------------
  #-----------------now i will work on details frame----------------------

  #here is little complicated so i will make a details_frame and inside that details_frame i will place  two more frames
  #                   Main frame   

details_frame = customtkinter.CTkFrame(master=root,
width=10, 
height=10)

  #as you can see master of details_frame is root which is main main window

#                     order detail frame

order_detail_frame = ttk.LabelFrame(master=details_frame,
   text= "Order Details",
   width=200,
   height=200)

##here I used the LabelFrame from ttk ( you will have to import this aswell) 


    #       here i will create entry boxes 

enter_product = customtkinter.CTkEntry(order_detail_frame,
    placeholder_text= "Enter product name",
    width=500,
    height=60,
    font=("helvetica",24),
    text_color="#F5DD90",
    placeholder_text_color="yellow",
    corner_radius=200)

#Quantity box
enter_quantity = customtkinter.CTkEntry(order_detail_frame,
    placeholder_text= "Enter Quantity",
    width=500,
    height=60,
    font=("helvetica",24),
    text_color="#F5DD90",
    placeholder_text_color="yellow",
    corner_radius=200)

    #as you can see i have these entry boxes in order detail frame

    #now i will create buttons


add_to_cart_button = customtkinter.CTkButton(details_frame,
    text ="add to cart",
    width =120,   # change width of button
    height =60,  # change height of button
    font=("helvetica",19),  # Change font of button
    text_color="black",  #  Change color of button
    fg_color="yellow",  # Change color of button foreground
    hover_color="#c2b84e",  #   Change color of when mouse hovers
    corner_radius=200,#    change corner raduis of button
    command=add_to_cart ) #and now we will add function to our button by calling def function created before



total_button = customtkinter.CTkButton(details_frame,
    text ="Total",
    width =120,
    height =60,
    font=("helvetica",19),
    text_color="black",
    fg_color="yellow",
    hover_color="#c2b84e",
    corner_radius=200,
    command= calculate_total)

    #as  you can see i have these buttons in details_frame outside of order_detail_frame




    #           customer  detail frame
customer_detail_frame = ttk.LabelFrame(master=details_frame,
   text= "Customer Details",
   width=200,
   height=200)


#same as i did with order_detail_frame

#and the master of these 2 frames is my details frame
customer_name = customtkinter.CTkEntry(customer_detail_frame,
    placeholder_text= "Customer Name",
    width=500,
    height=60,
    font=("helvetica",24),
    text_color="#F5DD90",
    placeholder_text_color="yellow",
    corner_radius=200) 



customer_number = customtkinter.CTkEntry(customer_detail_frame,
    placeholder_text= "Customer Number",
    width=500,
    height=60,
    font=("helvetica",24),
    text_color="#F5DD90",
    placeholder_text_color="yellow",
    corner_radius=200)

print_bill_button = customtkinter.CTkButton(details_frame, text ="Print Bill",
    width =120,
    height =60,
    font=("helvetica",19),
    text_color="black",
    fg_color="yellow",
    hover_color="#c2b84e",
    corner_radius=200 )

#now i will create a cart frame in which i will have a label cart and a listbox that i imported in starting

cart_frame = customtkinter.CTkFrame(master=root,
  width=5,
  height=5,
)


cart_label = customtkinter.CTkLabel(cart_frame,
    text ="Cart",
    font=('times new roman',30,'bold'),
  text_color='yellow',)


total_label = customtkinter.CTkLabel(cart_frame,
    text ="",
    font=('times new roman',30,'bold'),
  text_color='yellow',)


cart_listbox = CTkListbox(cart_frame,
  width=500,

  height=500,
  font=("helvetica"),)


#now i will create a bill frame in which i will have bill label and a text box which will print all bill




bill_frame = customtkinter.CTkFrame(master=root, width=200, height=50)

bill_label = customtkinter.CTkLabel(bill_frame,
    text ="Bill",
    font=('times new roman',30,'bold'),
  text_color='yellow',)

# here I will create a bill text box to put bill
bill_textbox = customtkinter.CTkTextbox(bill_frame,)


#here i created an  another frame for 3 button clear cart,save and 'exit'

cse_button_frame = customtkinter.CTkFrame(master=root, 
width=100,
height=200)



clear_button = customtkinter.CTkButton(cse_button_frame, text ="Clear Cart",
    width =120,
    height =60,
    font=("helvetica",19),
    text_color="#1B1B1E",
    fg_color="yellow",
    hover_color="#c2b84e",
    corner_radius=200,
    command=clear_cart)





exit_button = customtkinter.CTkButton(cse_button_frame, text ="Exit",
    width =120,
    height =60,
    font=("helvetica",19),
    text_color="black",
    fg_color="yellow",
    hover_color="#c2b84e",
    corner_radius=200,
    command=exit_program)





# Creating a grid for the widgets and frames

#firt i will make columns and rows with .configure
root.columnconfigure(0,weight =1,uniform="a")
root.columnconfigure(1,weight =1,uniform="a")
root.columnconfigure(2,weight =1,uniform="a")
root.rowconfigure(0,weight =1,uniform="a")
root.rowconfigure(1,weight =2,uniform="a")
root.rowconfigure(2,weight =1,uniform="a")

#giving weight is neccessary,weight means the size of the particular column

#now that i have configured the rows and column i will use.grid to place widegets in that row .grid takes few arguments like row,column,columnspan,sticky etc

#first i will grid the frames then widgets in them
#columnspan means how much columns our widgets will take
#sticky means direction
heading_frame.grid(column=0 ,row=0,columnspan=3, sticky=tkinter.NSEW, padx=5,pady=5)
details_frame.grid(column=0 ,row=1,rowspan=2, sticky=tkinter.NSEW, padx=5,pady=5)
cart_frame.grid(column=1 ,row=1, sticky=tkinter.NSEW,padx=75,pady=20)
bill_frame.grid(column=2 ,row=1, sticky=tkinter.NSEW, padx=75,pady=20)
cse_button_frame.grid(column=1 ,row=2,columnspan=2, sticky=tkinter.NSEW,padx=5,pady=5)

#now i will place widgets in frames with same grid method

#this time I will use the frame name instead of root

heading_frame.columnconfigure(0,weight =3,)
heading_frame.columnconfigure(1,weight =1,)
heading_frame.columnconfigure(2,weight =1,)
heading_frame.columnconfigure(3,weight =1,)
heading_frame.rowconfigure(0,weight =1)

headingLabel.grid(row=0, column=0,padx=5, sticky=tkinter.EW)

enter_username.grid(row=0, column=2,padx=5)

enter_password.grid(row=0, column=3,padx=5)

login_button.grid(row=0, column=4,padx=5)

#-------------heading frame done------------------


#-----------------now i will work on details frame----------------------

#here first I will make four rows and two column in my details frame
details_frame.rowconfigure(0,weight =1,uniform="a")
details_frame.rowconfigure(1,weight =1,uniform="a")
details_frame.rowconfigure(2,weight =1,uniform="a")
details_frame.rowconfigure(3,weight =1,uniform="a")
details_frame.columnconfigure(0,weight =1,uniform="a")
details_frame.columnconfigure(1,weight =1,uniform="a")

#now i will place widgets in details frame
#putting widgets in order detail frame

order_detail_frame.grid(row=0, column=0, columnspan=2,pady=10)
customer_detail_frame.grid(row=2, column=0,columnspan=2)

#configiring  order detail frame
order_detail_frame.columnconfigure(0,weight =1)
order_detail_frame.rowconfigure(0,weight =1)
order_detail_frame.rowconfigure(1,weight =1)

#configuring customer detail frame
customer_detail_frame.columnconfigure(0, weight=1,)
customer_detail_frame.rowconfigure(0, weight=1,)
customer_detail_frame.rowconfigure(1, weight=1,)

#griding widgets in order details frame

enter_product.grid(row=1, column=0,pady=10,padx=6)
enter_quantity.grid(row=2, column=0,pady=10,padx=6)

#griding button in  details frame

add_to_cart_button.grid(row=1, column=0,sticky=E,pady=10,padx=6)
total_button.grid(row=1, column=1,sticky=W,pady=10,padx=6)

#griding widgets in customer details frame
customer_name.grid(row=0, column=0,pady=10,padx=6)
customer_number.grid(row=1, column=0,pady=10,padx=6)
#griding button in customer details frame
print_bill_button.grid(row=3, column=0, columnspan=2)
#-----------------details frame done------------------

#-----------------now i will work on cart frame----------------------

#first i will make two rows and one columns in my cart frame
cart_frame.rowconfigure(0,weight =1)
cart_frame.rowconfigure(1,weight =5)
cart_frame.rowconfigure(2,weight =1)
cart_frame.columnconfigure(0,weight =1)
#now i will place widgets in cart frame
cart_label.grid(row=0, column=0,)
cart_listbox.grid(row=1, column=0,)
total_label.grid(row=2, column=0,)



#-----------------cart frame done------------------


#-----------------now i will work on bill box ------------------


bill_frame.rowconfigure(0,weight =1,uniform="a")
bill_frame.rowconfigure(1,weight =5,uniform="a")
bill_frame.columnconfigure(0,weight =1,uniform="a")
#now i will place widgets in cart frame
bill_label.grid(row=0, column=0,)
bill_textbox.grid(row=1, column=0,)


#now i will do in last cse buttons

cse_button_frame.columnconfigure(0,weight=1,uniform="a")
cse_button_frame.columnconfigure(1,weight=1,uniform="a")
cse_button_frame.columnconfigure(2,weight=1,uniform="a")
cse_button_frame.rowconfigure(0,weight=1)

clear_button.grid(row=0, column=0,padx=7, sticky=tkinter.EW)
exit_button.grid(row=0, column=1,padx=7, sticky=tkinter.EW)

# Start the main loop


root.mainloop()
#lets make a function that will show output from entry box when i click button

